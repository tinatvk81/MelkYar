

"""
سرویس Import اکسل — پیاده‌سازی کامل فاز ۲.

منطق کلی:
  1. فایل اکسل باید دقیقاً همان قالب ۴ شیتی باشد که قبلاً طراحی شد
     (شیت‌های: فروش، پیش‌خرید، اجاره، رهن_کامل).
  2. هر شیت سطر به سطر خوانده می‌شود؛ ستون فارسی هر سلول با نگاشت زیر به
     فیلد مدل جنگو تبدیل می‌شود.
  3. اگر سطری خطا داشت (مثلاً کد ملک تکراری، مشاور پیدا نشد، عدد نامعتبر)
     آن سطر رد می‌شود ولی بقیه‌ی فایل ادامه پیدا می‌کند — در پایان یک
     گزارش کامل {created, skipped, errors:[{row, sheet, message}]} برمی‌گردد.
  4. کل Import در یک تراکنش دیتابیسی نیست، چون می‌خواهیم سطرهای درست حتی
     اگر بعضی سطرها خطا داشتند ثبت شوند (تجربه‌ی بهتر برای وارد کردن حجم
     زیاد فایل به‌صورت یک‌جا).
"""
from datetime import date, datetime
import jdatetime

import openpyxl
from django.contrib.auth import get_user_model
from django.db import transaction

from properties.models import (
    MortgageDetail,
    PresaleDetail,
    Property,
    RentDetail,
    SaleDetail,
)

User = get_user_model()

YES_NO_MAP = {"بله": True, "خیر": False}

# نگاشت لیبل فارسی <-> مقدار enum مدل، برای فیلدهایی که choices دارند
PROPERTY_KIND_MAP = {label: key for key, label in Property.PropertyKind.choices}
DOCUMENT_TYPE_MAP = {label: key for key, label in Property.DocumentType.choices}
FILE_STATUS_MAP = {label: key for key, label in Property.FileStatus.choices}
PAYMENT_TERMS_MAP = {label: key for key, label in SaleDetail.PaymentTerms.choices}
RENEWAL_STATUS_MAP = {label: key for key, label in RentDetail.RenewalStatus.choices}

SHEET_TRANSACTION_TYPE = {
    "فروش": Property.TransactionType.SALE,
    "پیش‌خرید": Property.TransactionType.PRESALE,
    "اجاره": Property.TransactionType.RENT,
    "رهن_کامل": Property.TransactionType.MORTGAGE,
}

BASE_COLUMNS = [
    "کد ملک", "عنوان آگهی", "نوع ملک", "استان", "شهر", "منطقه/محله",
    "آدرس کامل", "لینک موقعیت (نقشه)", "متراژ زیربنا (متر)", "متراژ زمین (متر)",
    "تعداد اتاق خواب", "سال ساخت", "تعداد کل طبقات ساختمان", "طبقه واحد",
    "تعداد واحد در طبقه", "جهت ساختمان", "نوع سند", "آسانسور (بله/خیر)",
    "پارکینگ (تعداد)", "انباری (بله/خیر)", "بالکن (بله/خیر)",
    "امکانات رفاهی", "سیستم گرمایش", "سیستم سرمایش", "نوع کفپوش",
    "وضعیت بازسازی (بله/خیر)", "وضعیت واحد", "نام مالک", "شماره تماس مالک",
    "نوع فایل (انحصاری/غیرانحصاری)", "نام مشاور ثبت‌کننده", "تاریخ ثبت فایل",
    "تاریخ آخرین بروزرسانی", "وضعیت فایل", "توضیحات عمومی (نمایش به مشتری)",
    "یادداشت خصوصی مشاور", "پوشه/لینک تصاویر",
]

SALE_COLUMNS = ["قیمت کل (تومان)", "قیمت هر متر (تومان)", "شرایط پرداخت (نقد/اقساط)",
                "مبلغ پیش‌پرداخت", "قابل معاوضه (بله/خیر)"]
PRESALE_COLUMNS = ["قیمت کل قرارداد (تومان)", "شرکت/سازنده", "درصد پیشرفت پروژه",
                    "تاریخ تحویل تخمینی", "مبلغ پرداخت‌شده تاکنون", "مبلغ باقی‌مانده",
                    "شرایط اقساط باقی‌مانده", "شماره قرارداد پیش‌فروش"]
RENT_COLUMNS = ["مبلغ ودیعه/رهن (تومان)", "مبلغ اجاره ماهانه (تومان)",
                "تاریخ شروع قرارداد", "تاریخ پایان قرارداد", "روز تا پایان قرارداد",
                "نام مستاجر فعلی", "شماره تماس مستاجر فعلی",
                "قابل تبدیل رهن/اجاره (بله/خیر)", "درصد افزایش سالانه پیشنهادی",
                "وضعیت تمدید", "نتیجه آخرین تماس", "تاریخ تماس بعدی"]
MORTGAGE_COLUMNS = ["مبلغ رهن کامل (تومان)", "تاریخ شروع قرارداد", "تاریخ پایان قرارداد",
                     "روز تا پایان قرارداد", "نام مستاجر فعلی", "شماره تماس مستاجر فعلی",
                     "وضعیت تمدید", "نتیجه آخرین تماس", "تاریخ تماس بعدی"]


class RowError(Exception):
    pass


def _to_bool(value, field_label):
    if value in (None, ""):
        return False
    value = str(value).strip()
    if value not in YES_NO_MAP:
        raise RowError(f"مقدار «{field_label}» باید دقیقاً «بله» یا «خیر» باشد (مقدار داده‌شده: {value})")
    return YES_NO_MAP[value]


def _to_int(value, field_label, required=False):
    if value in (None, ""):
        if required:
            raise RowError(f"«{field_label}» نمی‌تواند خالی باشد.")
        return None
    try:
        return int(str(value).replace(",", "").replace("%", "").strip())
    except ValueError:
        raise RowError(f"«{field_label}» باید یک عدد باشد (مقدار داده‌شده: {value})")


PERSIAN_DIGITS = str.maketrans("۰۱۲۳۴۵۶۷۸۹", "0123456789")
ARABIC_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")


def _normalize_number_text(value):
    """
    تبدیل اعداد فارسی و عربی به انگلیسی.
    مثال:
        ۱۴۰۴/۰۸/۱۵ -> 1404/08/15
    """
    return str(value).translate(PERSIAN_DIGITS).translate(ARABIC_DIGITS)


def _parse_jalali_date(value, field_label):
    """
    تاریخ شمسی متنی را می‌پذیرد:
      1404/08/15
      ۱۴۰۴/۰۸/۱۵
      1404-08-15

    خروجی: datetime.date میلادی برای ذخیره در دیتابیس.
    """
    text = _normalize_number_text(value).strip()
    text = text.replace("-", "/").replace(".", "/")

    parts = [part.strip() for part in text.split("/") if part.strip()]
    if len(parts) != 3:
        raise RowError(
            f"«{field_label}» باید به فرم تاریخ شمسی "
            f"«1404/08/15» وارد شود (مقدار داده‌شده: {value})"
        )

    try:
        year, month, day = map(int, parts)

        # تاریخ‌های 13xx یا 14xx را شمسی می‌دانیم.
        if not 1200 <= year <= 1600:
            raise ValueError

        return jdatetime.date(year, month, day).togregorian()

    except (ValueError, TypeError):
        raise RowError(
            f"«{field_label}» تاریخ شمسی معتبر نیست "
            f"(مقدار داده‌شده: {value})"
        )


def _to_date(value, field_label, required=False):
    """
    تاریخ قابل قبول:
      - سلول Date اکسل: datetime/date میلادی
      - متن شمسی: 1404/08/15 یا ۱۴۰۴/۰۸/۱۵
      - متن ISO میلادی: 2026-11-06 (برای سازگاری با فایل‌های قبلی)
    """
    if value in (None, ""):
        if required:
            raise RowError(f"«{field_label}» نمی‌تواند خالی باشد.")
        return None

    # اگر اکسل یک Date واقعی ساخته باشد، openpyxl datetime/date می‌دهد.
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = _normalize_number_text(value).strip()

    # تاریخ شمسی متنی
    if "/" in text:
        return _parse_jalali_date(text, field_label)

    # پشتیبانی از تاریخ شمسی با -
    if "-" in text:
        parts = text.split("-")
        if len(parts) == 3 and len(parts[0]) == 4:
            try:
                year = int(parts[0])
                if 1200 <= year <= 1600:
                    return _parse_jalali_date(text, field_label)
            except ValueError:
                pass

        # برای سازگاری با تاریخ میلادی ISO مثل 2026-08-01
        try:
            return datetime.strptime(text, "%Y-%m-%d").date()
        except ValueError:
            pass

    raise RowError(
        f"«{field_label}» باید تاریخ شمسی معتبر مانند «1404/08/15» باشد "
        f"(مقدار داده‌شده: {value})"
    )


def _map_choice(value, mapping, field_label, required=True):
    if value in (None, ""):
        if required:
            raise RowError(f"«{field_label}» نمی‌تواند خالی باشد.")
        return ""
    value = str(value).strip()
    if value not in mapping:
        valid = "، ".join(mapping.keys())
        raise RowError(f"«{field_label}» نامعتبر است: «{value}». مقادیر مجاز: {valid}")
    return mapping[value]


def _row_dict(header_row, row):
    return {header_row[i]: row[i] for i in range(len(header_row)) if i < len(row)}


def _build_base_fields(data, agent_cache, errors_prefix=""):
    agent_name = str(data.get("نام مشاور ثبت‌کننده") or "").strip()
    if not agent_name:
        raise RowError("«نام مشاور ثبت‌کننده» نمی‌تواند خالی باشد.")
    if agent_name not in agent_cache:
        raise RowError(
            f"مشاوری با نام «{agent_name}» در سیستم پیدا نشد. اول از پنل مدیریت کاربران اکانتش را بسازید."
        )

    code = str(data.get("کد ملک") or "").strip()
    if not code:
        raise RowError("«کد ملک» نمی‌تواند خالی باشد.")
    if Property.objects.filter(code=code).exists():
        raise RowError(f"کد ملک «{code}» قبلاً در سیستم ثبت شده (تکراری).")

    return {
        "owner_agent": agent_cache[agent_name],
        "code": code,
        "title": str(data.get("عنوان آگهی") or "").strip(),
        "property_kind": _map_choice(data.get("نوع ملک"), PROPERTY_KIND_MAP, "نوع ملک"),
        "province": str(data.get("استان") or "").strip(),
        "city": str(data.get("شهر") or "").strip(),
        "district": str(data.get("منطقه/محله") or "").strip(),
        "full_address": str(data.get("آدرس کامل") or "").strip(),
        "map_link": str(data.get("لینک موقعیت (نقشه)") or "").strip(),
        "area_sqm": _to_int(data.get("متراژ زیربنا (متر)"), "متراژ زیربنا"),
        "land_area_sqm": _to_int(data.get("متراژ زمین (متر)"), "متراژ زمین"),
        "bedrooms": _to_int(data.get("تعداد اتاق خواب"), "تعداد اتاق خواب"),
        "build_year": _to_int(data.get("سال ساخت"), "سال ساخت"),
        "total_floors": _to_int(data.get("تعداد کل طبقات ساختمان"), "تعداد کل طبقات"),
        "unit_floor": _to_int(data.get("طبقه واحد"), "طبقه واحد"),
        "units_per_floor": _to_int(data.get("تعداد واحد در طبقه"), "تعداد واحد در طبقه"),
        "direction": str(data.get("جهت ساختمان") or "").strip(),
        "document_type": _map_choice(data.get("نوع سند"), DOCUMENT_TYPE_MAP, "نوع سند", required=False),
        "has_elevator": _to_bool(data.get("آسانسور (بله/خیر)"), "آسانسور"),
        "parking_count": _to_int(data.get("پارکینگ (تعداد)"), "پارکینگ") or 0,
        "has_storage": _to_bool(data.get("انباری (بله/خیر)"), "انباری"),
        "has_balcony": _to_bool(data.get("بالکن (بله/خیر)"), "بالکن"),
        "amenities": str(data.get("امکانات رفاهی") or "").strip(),
        "heating_system": str(data.get("سیستم گرمایش") or "").strip(),
        "cooling_system": str(data.get("سیستم سرمایش") or "").strip(),
        "floor_covering": str(data.get("نوع کفپوش") or "").strip(),
        "is_renovated": _to_bool(data.get("وضعیت بازسازی (بله/خیر)"), "وضعیت بازسازی"),
        "unit_condition": str(data.get("وضعیت واحد") or "").strip(),
        "owner_name": str(data.get("نام مالک") or "").strip(),
        "owner_phone": str(data.get("شماره تماس مالک") or "").strip(),
        "is_exclusive": str(data.get("نوع فایل (انحصاری/غیرانحصاری)") or "").strip() == "انحصاری",
        "status": _map_choice(data.get("وضعیت فایل"), FILE_STATUS_MAP, "وضعیت فایل", required=False) or Property.FileStatus.ACTIVE,
        "public_description": str(data.get("توضیحات عمومی (نمایش به مشتری)") or "").strip(),
        "private_note": str(data.get("یادداشت خصوصی مشاور") or "").strip(),
    }


def _build_sale_detail(data):
    return {
        "total_price": _to_int(data.get("قیمت کل (تومان)"), "قیمت کل", required=True),
        "price_per_sqm": _to_int(data.get("قیمت هر متر (تومان)"), "قیمت هر متر"),
        "payment_terms": _map_choice(data.get("شرایط پرداخت (نقد/اقساط)"), PAYMENT_TERMS_MAP, "شرایط پرداخت", required=False) or SaleDetail.PaymentTerms.CASH,
        "down_payment": _to_int(data.get("مبلغ پیش‌پرداخت"), "مبلغ پیش‌پرداخت"),
        "is_exchangeable": _to_bool(data.get("قابل معاوضه (بله/خیر)"), "قابل معاوضه"),
    }


def _build_presale_detail(data):
    return {
        "total_contract_price": _to_int(data.get("قیمت کل قرارداد (تومان)"), "قیمت کل قرارداد", required=True),
        "builder_company": str(data.get("شرکت/سازنده") or "").strip(),
        "progress_percent": _to_int(data.get("درصد پیشرفت پروژه"), "درصد پیشرفت پروژه"),
        "estimated_delivery_date": _to_date(data.get("تاریخ تحویل تخمینی"), "تاریخ تحویل تخمینی"),
        "amount_paid": _to_int(data.get("مبلغ پرداخت‌شده تاکنون"), "مبلغ پرداخت‌شده"),
        "amount_remaining": _to_int(data.get("مبلغ باقی‌مانده"), "مبلغ باقی‌مانده"),
        "installment_terms": str(data.get("شرایط اقساط باقی‌مانده") or "").strip(),
        "contract_number": str(data.get("شماره قرارداد پیش‌فروش") or "").strip(),
    }


def _build_rent_detail(data):
    return {
        "deposit_amount": _to_int(data.get("مبلغ ودیعه/رهن (تومان)"), "مبلغ ودیعه", required=True),
        "monthly_rent": _to_int(data.get("مبلغ اجاره ماهانه (تومان)"), "اجاره ماهانه", required=True),
        "contract_start_date": _to_date(data.get("تاریخ شروع قرارداد"), "تاریخ شروع قرارداد", required=True),
        "contract_end_date": _to_date(data.get("تاریخ پایان قرارداد"), "تاریخ پایان قرارداد", required=True),
        "current_tenant_name": str(data.get("نام مستاجر فعلی") or "").strip(),
        "current_tenant_phone": str(data.get("شماره تماس مستاجر فعلی") or "").strip(),
        "convertible_to_mortgage": _to_bool(data.get("قابل تبدیل رهن/اجاره (بله/خیر)"), "قابل تبدیل"),
        "yearly_increase_percent": _to_int(data.get("درصد افزایش سالانه پیشنهادی"), "درصد افزایش سالانه"),
        "renewal_status": _map_choice(data.get("وضعیت تمدید"), RENEWAL_STATUS_MAP, "وضعیت تمدید", required=False) or RentDetail.RenewalStatus.NOT_CHECKED,
        "last_contact_result": str(data.get("نتیجه آخرین تماس") or "").strip(),
        "next_contact_date": _to_date(data.get("تاریخ تماس بعدی"), "تاریخ تماس بعدی"),
    }


def _build_mortgage_detail(data):
    return {
        "deposit_amount": _to_int(data.get("مبلغ رهن کامل (تومان)"), "مبلغ رهن کامل", required=True),
        "contract_start_date": _to_date(data.get("تاریخ شروع قرارداد"), "تاریخ شروع قرارداد", required=True),
        "contract_end_date": _to_date(data.get("تاریخ پایان قرارداد"), "تاریخ پایان قرارداد", required=True),
        "current_tenant_name": str(data.get("نام مستاجر فعلی") or "").strip(),
        "current_tenant_phone": str(data.get("شماره تماس مستاجر فعلی") or "").strip(),
        "renewal_status": _map_choice(data.get("وضعیت تمدید"), RENEWAL_STATUS_MAP, "وضعیت تمدید", required=False) or RentDetail.RenewalStatus.NOT_CHECKED,
        "last_contact_result": str(data.get("نتیجه آخرین تماس") or "").strip(),
        "next_contact_date": _to_date(data.get("تاریخ تماس بعدی"), "تاریخ تماس بعدی"),
    }


SHEET_HANDLERS = {
    "فروش": (SaleDetail, _build_sale_detail, "sale_detail"),
    "پیش‌خرید": (PresaleDetail, _build_presale_detail, "presale_detail"),
    "اجاره": (RentDetail, _build_rent_detail, "rent_detail"),
    "رهن_کامل": (MortgageDetail, _build_mortgage_detail, "mortgage_detail"),
}


def import_excel_file(file_obj):
    """
    ورودی: فایل آپلودشده (InMemoryUploadedFile یا مسیر).
    خروجی: dict گزارش -> {"created": int, "skipped": int, "errors": [...]}
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    agent_cache = {}
    for user in User.objects.filter(role=User.Role.AGENT) | User.objects.filter(role=User.Role.ADMIN):
        agent_cache[user.get_full_name() or user.username] = user
        agent_cache[user.username] = user

    report = {"created": 0, "skipped": 0, "errors": []}

    for sheet_name, transaction_type in SHEET_TRANSACTION_TYPE.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header_row = rows[0]

        detail_model, build_detail_fn, rel_name = SHEET_HANDLERS[sheet_name]

        for row_idx, row in enumerate(rows[1:], start=2):
            if row is None or all(cell in (None, "") for cell in row):
                continue  # سطر کاملاً خالی، رد شود

            data = _row_dict(header_row, row)
            try:
                with transaction.atomic():
                    base_fields = _build_base_fields(data, agent_cache)
                    base_fields["transaction_type"] = transaction_type
                    detail_fields = build_detail_fn(data)

                    property_obj = Property.objects.create(**base_fields)
                    detail_model.objects.create(property=property_obj, **detail_fields)

                report["created"] += 1
            except RowError as exc:
                report["skipped"] += 1
                report["errors"].append({"sheet": sheet_name, "row": row_idx, "message": str(exc)})
            except Exception as exc:  # خطای غیرمنتظره؛ سطر رد می‌شود ولی importها متوقف نمی‌شود
                report["skipped"] += 1
                report["errors"].append({"sheet": sheet_name, "row": row_idx, "message": f"خطای غیرمنتظره: {exc}"})

    return report
